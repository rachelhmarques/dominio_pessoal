from __future__ import annotations

import csv
import json
import re
import struct
import subprocess
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Iterable
from uuid import uuid4

import olefile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CSV_HEADER = [
    "Data",
    "Cód. Conta Debito",
    "Cód. Conta Credito",
    "Valor",
    "Cód. Histórico",
    "Complemento Histórico",
    "Inicia lote",
    "Código Matriz/Filial",
    "Centro de Custo Débito",
    "Centro de Custo Crédito",
]

TWO_PLACES = Decimal("0.01")
SOURCE_DEFINITIONS = {
    "vacation_total": {
        "label": "Provisão de férias",
        "description": "1/3 DAS FERIAS + DIAS FERIAS",
        "memory_prefix": "Soma das rubricas de férias",
        "order": 10,
    },
    "thirteenth_difference": {
        "label": "Diferença de 13º",
        "description": "Rubrica DIFERENCA 13o",
        "memory_prefix": "Valor da diferença de 13º",
        "order": 20,
    },
    "regular_salary": {
        "label": "Salários regulares",
        "description": "Soma dos proventos elegíveis, excluindo férias e diferença de 13º",
        "memory_prefix": "Soma dos proventos elegíveis",
        "order": 30,
    },
    "vacation_advance": {
        "label": "Adiantamento de férias",
        "description": "Rubrica ADIANTAMENTO DE FERIAS",
        "memory_prefix": "Valor do adiantamento de férias",
        "order": 40,
    },
    "transport_discount": {
        "label": "Vale-transporte descontado em folha",
        "description": "Rubrica VALE TRANSPORTE",
        "memory_prefix": "Valor do vale-transporte descontado em folha",
        "order": 45,
    },
    "net_salary": {
        "label": "Salários a pagar",
        "description": "Campo Líquido Geral",
        "memory_prefix": "Valor líquido da folha",
        "order": 50,
    },
    "employee_inss": {
        "label": "INSS empregado",
        "description": "Campo Segurados",
        "memory_prefix": "INSS descontado do empregado",
        "order": 60,
    },
    "fgts": {
        "label": "FGTS",
        "description": "Campo Valor do FGTS",
        "memory_prefix": "FGTS calculado no resumo",
        "order": 70,
    },
    "pis": {
        "label": "PIS",
        "description": "Campo Valor PIS",
        "memory_prefix": "PIS calculado no resumo",
        "order": 80,
    },
    "employer_inss": {
        "label": "INSS patronal",
        "description": "Empresa + RAT + Terceiros",
        "memory_prefix": "Soma dos encargos patronais",
        "order": 90,
    },
    "thirteenth_salary": {
        "label": "13Âº integral",
        "description": "Proventos do Resumo Mensal13",
        "memory_prefix": "Soma dos proventos do 13Âº",
        "order": 100,
    },
    "thirteenth_net_salary": {
        "label": "13Âº a pagar",
        "description": "Campo LÃ­quido Geral do Resumo Mensal13",
        "memory_prefix": "Valor lÃ­quido do 13Âº",
        "order": 110,
    },
    "thirteenth_advance": {
        "label": "Adiantamento 13Âº",
        "description": "Rubrica ADIANTAMENTO 13 SALARIO",
        "memory_prefix": "Valor do adiantamento do 13Âº",
        "order": 120,
    },
    "thirteenth_employee_inss": {
        "label": "INSS 13Âº empregado",
        "description": "Campo Segurados do Resumo Mensal13",
        "memory_prefix": "INSS descontado no 13Âº",
        "order": 130,
    },
    "thirteenth_inss_discounts": {
        "label": "INSS descontos 13Ã‚Âº",
        "description": "SalÃ¡rio-famÃ­lia e salÃ¡rio-maternidade compensÃ¡veis no Resumo Mensal13",
        "memory_prefix": "Soma dos descontos compensÃ¡veis do INSS no 13Ã‚Âº",
        "order": 135,
    },
    "thirteenth_irrf": {
        "label": "IRRF 13Ã‚Âº",
        "description": "Campo Valor IRRF 13Ã‚Âº SalÃ¡rio do Resumo Mensal13",
        "memory_prefix": "IRRF descontado no 13Ã‚Âº",
        "order": 137,
    },
    "thirteenth_fgts": {
        "label": "FGTS 13Âº",
        "description": "Campo Valor do FGTS do Resumo Mensal13",
        "memory_prefix": "FGTS calculado no 13Âº",
        "order": 140,
    },
    "thirteenth_pis": {
        "label": "PIS 13Âº",
        "description": "Campo Valor PIS do Resumo Mensal13",
        "memory_prefix": "PIS calculado no 13Âº",
        "order": 150,
    },
    "thirteenth_employer_inss": {
        "label": "INSS 13Âº patronal",
        "description": "Empresa + RAT + Terceiros do Resumo Mensal13",
        "memory_prefix": "Soma dos encargos patronais do 13Âº",
        "order": 160,
    },
}
SOURCE_DEFINITIONS.update(
    {
        "inss_discounts": {
            "label": "INSS descontos",
            "description": "Proventos compensáveis em INSS, como salário-família, maternidade e afastamentos elegíveis",
            "memory_prefix": "Soma dos proventos compensáveis em INSS",
            "order": 35,
        },
        "health_plan_reimbursement": {
            "label": "Reembolso plano de saúde",
            "description": "Rubrica REEMBOLSO PLANO DE SAUDE",
            "memory_prefix": "Valor do reembolso de plano de saúde",
            "order": 37,
        },
        "termination_net": {
            "label": "Rescisão líquida",
            "description": "Rubrica LIQUIDO RESCISAO",
            "memory_prefix": "Valor líquido das rescisões",
            "order": 47,
        },
        "absence_adjustment": {
            "label": "Afastamentos e faltas",
            "description": "Rubricas de descontos por afastamentos e faltas",
            "memory_prefix": "Soma dos descontos por afastamentos e faltas",
            "order": 48,
        },
        "employee_irrf": {
            "label": "IRRF empregados",
            "description": "IRRF mensal, férias e rescisão de empregados",
            "memory_prefix": "Soma do IRRF de empregados",
            "order": 55,
        },
        "health_plan_discount": {
            "label": "Desconto plano de saúde",
            "description": "Rubrica DESCONTO PLANO DE SAÚDE",
            "memory_prefix": "Valor descontado de plano de saúde",
            "order": 85,
        },
        "meal_voucher_discount": {
            "label": "Vale alimentaÃ§Ã£o",
            "description": "Rubrica DESC VALE ALIMENTACAO",
            "memory_prefix": "Valor descontado de vale alimentaÃ§Ã£o",
            "order": 86,
        },
        "alimony": {
            "label": "PensÃ£o alimentÃ­cia",
            "description": "Rubrica PENSAO ALIMENTICIA",
            "memory_prefix": "Valor da pensÃ£o alimentÃ­cia descontada em folha",
            "order": 87,
        },
        "termination_charge_debit": {
            "label": "Encargos de rescisão",
            "description": "Aviso prévio, estouro de rescisão e multa de estabilidade do empregador",
            "memory_prefix": "Soma dos encargos patronais de rescisão",
            "order": 46,
        },
        "termination_charge_credit": {
            "label": "Multa rescisão",
            "description": "Rubrica MULTA ESTABILIDADE Art. 480/CLT",
            "memory_prefix": "Valor da multa de rescisão descontada",
            "order": 46,
        },
        "prior_notice_recovery": {
            "label": "Aviso prévio reavido",
            "description": "Rubrica AVISO PREVIO REAVIDO",
            "memory_prefix": "Valor do aviso prévio reavido",
            "order": 46,
        },
        "thirteenth_rescisao_discount": {
            "label": "13º salário rescisão",
            "description": "Adiantamento e descontos de diferença de 13º na rescisão",
            "memory_prefix": "Soma do 13º descontado na rescisão",
            "order": 121,
        },
        "payroll_discount": {
            "label": "Desconto folha",
            "description": "Descontos de crédito do trabalhador e adiantamentos salariais",
            "memory_prefix": "Soma dos descontos lançados diretamente na folha",
            "order": 49,
        },
    }
)
THIRTEENTH_SUMMARY_SOURCE_KEYS = {
    "thirteenth_salary",
    "thirteenth_net_salary",
    "thirteenth_advance",
    "thirteenth_employee_inss",
    "thirteenth_inss_discounts",
    "thirteenth_irrf",
    "thirteenth_fgts",
    "thirteenth_pis",
    "thirteenth_employer_inss",
}
START_LOT_OPTIONS = {
    "always": "Sempre",
    "never": "Nunca",
    "debit": "Se houver débito",
    "no_prior_specials": "Somente se não houver provisão anterior",
}


@dataclass
class WorksheetRow:
    sheet: str
    row_number: int
    cells: list[str]


@dataclass
class CompetencyBlock:
    competency_date: date
    rows: list[WorksheetRow]


@dataclass
class MappingRule:
    rule_id: str
    label: str
    source_key: str
    debit_account: str
    credit_account: str
    history_template: str
    start_lot_strategy: str
    active: bool
    order: int

    def to_dict(self) -> dict[str, str | bool | int]:
        return {
            "rule_id": self.rule_id,
            "label": self.label,
            "source_key": self.source_key,
            "debit_account": self.debit_account,
            "credit_account": self.credit_account,
            "history_template": self.history_template,
            "start_lot_strategy": self.start_lot_strategy,
            "active": self.active,
            "order": self.order,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, str | bool | int]) -> "MappingRule":
        return cls(
            rule_id=str(payload["rule_id"]),
            label=str(payload["label"]),
            source_key=str(payload["source_key"]),
            debit_account=str(payload["debit_account"]),
            credit_account=str(payload["credit_account"]),
            history_template=str(payload["history_template"]),
            start_lot_strategy=str(payload["start_lot_strategy"]),
            active=bool(payload["active"]),
            order=int(payload["order"]),
        )


@dataclass
class CalculationContext:
    competency_date: date
    posting_date: str
    reference: str
    values: dict[str, Decimal]
    components: dict[str, list[tuple[str, Decimal]]]
    flags: dict[str, bool]

    def to_state(self) -> dict[str, object]:
        return {
            "competency_date": self.competency_date.isoformat(),
            "posting_date": self.posting_date,
            "reference": self.reference,
            "values": {key: decimal_to_storage(value) for key, value in self.values.items()},
            "components": {
                key: [{"label": label, "value": decimal_to_storage(value)} for label, value in items]
                for key, items in self.components.items()
            },
            "flags": self.flags,
        }

    @classmethod
    def from_state(cls, payload: dict[str, object]) -> "CalculationContext":
        raw_components = payload["components"]
        return cls(
            competency_date=date.fromisoformat(str(payload["competency_date"])),
            posting_date=str(payload["posting_date"]),
            reference=str(payload["reference"]),
            values={
                key: storage_to_decimal(value)
                for key, value in dict(payload["values"]).items()
            },
            components={
                key: [
                    (str(item["label"]), storage_to_decimal(item["value"]))
                    for item in items
                ]
                for key, items in dict(raw_components).items()
            },
            flags={key: bool(value) for key, value in dict(payload["flags"]).items()},
        )


@dataclass
class GeneratedEntry:
    competency_reference: str
    mapping_label: str
    source_key: str
    source_label: str
    posting_date: str
    debit_account: str
    credit_account: str
    value: Decimal
    history: str
    start_lot: bool
    memory: str
    branch_code: str
    order: int

    def to_csv_row(self) -> dict[str, str]:
        return {
            "Data": self.posting_date,
            "Cód. Conta Debito": self.debit_account,
            "Cód. Conta Credito": self.credit_account,
            "Valor": format_value(self.value),
            "Cód. Histórico": "",
            "Complemento Histórico": self.history,
            "Inicia lote": "1" if self.start_lot else "",
            "Código Matriz/Filial": self.branch_code,
            "Centro de Custo Débito": "",
            "Centro de Custo Crédito": "",
        }


class DomainParsingError(RuntimeError):
    pass


class SafeFormatDict(dict[str, str]):
    def __missing__(self, key: str) -> str:
        return "{" + key + "}"


def normalized_history_template(rule: MappingRule) -> str:
    debit = rule.debit_account.strip()
    credit = rule.credit_account.strip()

    if rule.source_key in {"employee_inss", "employer_inss"}:
        if debit and credit:
            return "INSS REF {reference}"
        if credit and not debit:
            return "INSS A RECOLHER {reference}"

    if rule.source_key in {"fgts", "thirteenth_fgts"}:
        if debit and credit:
            return "FGTS REF {reference}"
        if credit and not debit:
            return "FGTS A RECOLHER {reference}"

    if rule.source_key in {"pis", "thirteenth_pis"}:
        if debit and credit:
            return "PIS REF {reference}"
        if credit and not debit:
            return "PIS A RECOLHER {reference}"

    return rule.history_template


def apply_history_template_conventions(rules: list[MappingRule]) -> list[MappingRule]:
    for rule in rules:
        rule.history_template = sanitize_text(normalized_history_template(rule))
    return rules


def default_mapping_rules(
    branch_code: str | None = None,
    include_thirteenth_summary: bool = False,
) -> list[MappingRule]:
    rules = [
        MappingRule(
            rule_id="vacation_provision",
            label="Provisão férias",
            source_key="vacation_total",
            debit_account="1924",
            credit_account="",
            history_template="PROVISAO FERIAS {reference}",
            start_lot_strategy="always",
            active=True,
            order=10,
        ),
        MappingRule(
            rule_id="thirteenth_provision",
            label="Provisão 13º",
            source_key="thirteenth_difference",
            debit_account="1923",
            credit_account="",
            history_template="PROVISAO 13º {reference}",
            start_lot_strategy="always",
            active=True,
            order=20,
        ),
        MappingRule(
            rule_id="regular_salary",
            label="Salários",
            source_key="regular_salary",
            debit_account="2583",
            credit_account="",
            history_template="SALARIOS {reference}",
            start_lot_strategy="no_prior_specials",
            active=True,
            order=30,
        ),
        MappingRule(
            rule_id="vacation_advance",
            label="Adiantamento de férias",
            source_key="vacation_advance",
            debit_account="",
            credit_account="1764",
            history_template="ADIANTAMENTO FERIAS {reference}",
            start_lot_strategy="never",
            active=True,
            order=40,
        ),
        MappingRule(
            rule_id="transport_discount",
            label="Vale-transporte descontado em folha",
            source_key="transport_discount",
            debit_account="",
            credit_account="2806",
            history_template="VALE TRANSPORTE {reference}",
            start_lot_strategy="never",
            active=True,
            order=45,
        ),
        MappingRule(
            rule_id="net_salary",
            label="Salários a pagar",
            source_key="net_salary",
            debit_account="",
            credit_account="1902",
            history_template="SALARIOS A PAGAR {reference}",
            start_lot_strategy="never",
            active=True,
            order=50,
        ),
        MappingRule(
            rule_id="employee_inss",
            label="INSS empregado",
            source_key="employee_inss",
            debit_account="",
            credit_account="1903",
            history_template="INSS A RECOLHER {reference}",
            start_lot_strategy="never",
            active=True,
            order=60,
        ),
        MappingRule(
            rule_id="fgts",
            label="FGTS",
            source_key="fgts",
            debit_account="2587",
            credit_account="1904",
            history_template="FGTS REF {reference}",
            start_lot_strategy="always",
            active=True,
            order=70,
        ),
        MappingRule(
            rule_id="pis",
            label="PIS",
            source_key="pis",
            debit_account="2593",
            credit_account="1915",
            history_template="PIS REF {reference}",
            start_lot_strategy="always",
            active=True,
            order=80,
        ),
        MappingRule(
            rule_id="employer_inss",
            label="INSS patronal",
            source_key="employer_inss",
            debit_account="2586",
            credit_account="1903",
            history_template="INSS REF {reference}",
            start_lot_strategy="always",
            active=True,
            order=90,
        ),
    ]
    if branch_code == "67":
        for rule in rules:
            if rule.rule_id == "thirteenth_provision":
                rule.start_lot_strategy = "never"
            if rule.rule_id == "vacation_advance":
                rule.order = 41
            if rule.rule_id == "transport_discount":
                rule.history_template = "DESCONTO S FOLHA {reference}"
                rule.order = 39
        rules.extend(
            [
                MappingRule(
                    rule_id="inss_discounts",
                    label="INSS descontos",
                    source_key="inss_discounts",
                    debit_account="1903",
                    credit_account="",
                    history_template="INSS DESCONTOS {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=35,
                ),
                MappingRule(
                    rule_id="health_plan_reimbursement",
                    label="Reembolso plano de saúde",
                    source_key="health_plan_reimbursement",
                    debit_account="2641",
                    credit_account="",
                    history_template="REEMBOLSO PLANO DE SAUDE {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=37,
                ),
                MappingRule(
                    rule_id="health_plan_discount",
                    label="Desconto plano de saúde",
                    source_key="health_plan_discount",
                    debit_account="",
                    credit_account="2641",
                    history_template="DESCONTO PLANO DE SAUDE {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=38,
                ),
                MappingRule(
                    rule_id="termination_net",
                    label="Rescisão",
                    source_key="termination_net",
                    debit_account="",
                    credit_account="1906",
                    history_template="RESCISAO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=40,
                ),
                MappingRule(
                    rule_id="absence_adjustment",
                    label="Afastamento",
                    source_key="absence_adjustment",
                    debit_account="",
                    credit_account="2583",
                    history_template="AFASTAMENTO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=42,
                ),
                MappingRule(
                    rule_id="employee_irrf",
                    label="IRRF empregados",
                    source_key="employee_irrf",
                    debit_account="",
                    credit_account="1909",
                    history_template="IRRF EMPREGADOS {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=58,
                ),
                MappingRule(
                    rule_id="meal_voucher_discount",
                    label="Vale alimentação",
                    source_key="meal_voucher_discount",
                    debit_account="",
                    credit_account="2607",
                    history_template="VALE ALIMENTACAO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=59,
                ),
                MappingRule(
                    rule_id="alimony",
                    label="Pensão alimentícia",
                    source_key="alimony",
                    debit_account="",
                    credit_account="3705",
                    history_template="PENSAO ALIMENTICIA {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=59,
                ),
                MappingRule(
                    rule_id="termination_charge_debit",
                    label="Encargos rescisão débito",
                    source_key="termination_charge_debit",
                    debit_account="2588",
                    credit_account="",
                    history_template="MULTA RESCISAO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=36,
                ),
                MappingRule(
                    rule_id="prior_notice_recovery",
                    label="Aviso prévio reavido",
                    source_key="prior_notice_recovery",
                    debit_account="",
                    credit_account="2588",
                    history_template="AVISO PREVIO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=36,
                ),
                MappingRule(
                    rule_id="termination_charge_credit",
                    label="Multa rescisão crédito",
                    source_key="termination_charge_credit",
                    debit_account="",
                    credit_account="2588",
                    history_template="MULTA RESCISAO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=43,
                ),
                MappingRule(
                    rule_id="thirteenth_rescisao_discount",
                    label="13º salário rescisão",
                    source_key="thirteenth_rescisao_discount",
                    debit_account="",
                    credit_account="1763",
                    history_template="13ª SALARIO {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=57,
                ),
                MappingRule(
                    rule_id="payroll_discount",
                    label="Desconto folha",
                    source_key="payroll_discount",
                    debit_account="",
                    credit_account="2583",
                    history_template="DESCONTO FOLHA {reference}",
                    start_lot_strategy="never",
                    active=True,
                    order=44,
                ),
            ]
        )
    if include_thirteenth_summary:
        rules.extend(
            [
                MappingRule(
                    rule_id="thirteenth_salary",
                    label="13Âº integral",
                    source_key="thirteenth_salary",
                    debit_account="1923",
                    credit_account="",
                    history_template="PROVISAO 13Âª SALARIO",
                    start_lot_strategy="always",
                    active=True,
                    order=100,
                ),
                MappingRule(
                    rule_id="thirteenth_net_salary",
                    label="13Âº a pagar",
                    source_key="thirteenth_net_salary",
                    debit_account="",
                    credit_account="1910",
                    history_template="13Âª A PAGAR",
                    start_lot_strategy="never",
                    active=True,
                    order=110,
                ),
                MappingRule(
                    rule_id="thirteenth_advance",
                    label="Adiantamento 13Âº",
                    source_key="thirteenth_advance",
                    debit_account="",
                    credit_account="1763",
                    history_template="ADIANTAMENTO 13Âª SALARIO",
                    start_lot_strategy="never",
                    active=True,
                    order=120,
                ),
                MappingRule(
                    rule_id="thirteenth_employee_inss",
                    label="INSS 13Âº empregado",
                    source_key="thirteenth_employee_inss",
                    debit_account="",
                    credit_account="1903",
                    history_template="INSS 13Âª SALARIO",
                    start_lot_strategy="never",
                    active=True,
                    order=130,
                ),
                MappingRule(
                    rule_id="thirteenth_inss_discounts",
                    label="INSS descontos 13Ã‚Âº",
                    source_key="thirteenth_inss_discounts",
                    debit_account="1903",
                    credit_account="",
                    history_template="INSS DESCONTOS ",
                    start_lot_strategy="never",
                    active=True,
                    order=135,
                ),
                MappingRule(
                    rule_id="thirteenth_irrf",
                    label="IRRF 13Ã‚Âº",
                    source_key="thirteenth_irrf",
                    debit_account="",
                    credit_account="1909",
                    history_template="IRR 13Ã‚Âª SALARIO",
                    start_lot_strategy="never",
                    active=True,
                    order=137,
                ),
                MappingRule(
                    rule_id="thirteenth_fgts",
                    label="FGTS 13Âº",
                    source_key="thirteenth_fgts",
                    debit_account="2587",
                    credit_account="1904",
                    history_template="FGTS REF 13Âª SALARIO",
                    start_lot_strategy="always",
                    active=True,
                    order=140,
                ),
                MappingRule(
                    rule_id="thirteenth_pis",
                    label="PIS 13Âº",
                    source_key="thirteenth_pis",
                    debit_account="2593",
                    credit_account="1915",
                    history_template="PIS REF 13? SALARIO",
                    start_lot_strategy="always",
                    active=True,
                    order=150,
                ),
                MappingRule(
                    rule_id="thirteenth_employer_inss",
                    label="INSS 13Âº patronal",
                    source_key="thirteenth_employer_inss",
                    debit_account="2586",
                    credit_account="1903",
                    history_template="INSS REF 13? SALARIO",
                    start_lot_strategy="always",
                    active=True,
                    order=160,
                ),
            ]
        )
        for rule in rules:
            if rule.source_key in THIRTEENTH_SUMMARY_SOURCE_KEYS:
                rule.label = rule.label.replace("Âº", "\u00ba")
                rule.history_template = rule.history_template.replace("Âª", "\u00aa")
    for rule in rules:
        rule.label = sanitize_text(rule.label)
        rule.history_template = sanitize_text(rule.history_template)
    return apply_history_template_conventions(rules)


def build_empty_mapping_rule(next_order: int | None = None) -> MappingRule:
    return MappingRule(
        rule_id=f"custom_{uuid4().hex[:8]}",
        label="Novo mapeamento",
        source_key="regular_salary",
        debit_account="",
        credit_account="",
        history_template="NOVO HISTORICO {reference}",
        start_lot_strategy="never",
        active=True,
        order=next_order if next_order is not None else 999,
    )


def run_excel_extractor(workbook_path: str | Path) -> list[WorksheetRow]:
    workbook_path = Path(workbook_path)
    if workbook_path.suffix.lower() == ".xlsx":
        return run_openpyxl_extractor(workbook_path)
    try:
        return run_biff_extractor(workbook_path)
    except Exception:
        return run_powershell_extractor(workbook_path)


def run_openpyxl_extractor(workbook_path: Path) -> list[WorksheetRow]:
    workbook = load_workbook(filename=str(workbook_path), data_only=True)
    rows: list[WorksheetRow] = []

    for worksheet in workbook.worksheets:
        for row_number, excel_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            cells = []
            for cell_value in excel_row:
                text = openpyxl_value_to_text(cell_value)
                if text:
                    cells.append(text)
            if cells:
                rows.append(
                    WorksheetRow(
                        sheet=worksheet.title,
                        row_number=row_number,
                        cells=cells,
                    )
                )

    if not rows:
        raise DomainParsingError("Nenhuma linha útil foi encontrada na planilha.")
    return rows


def openpyxl_value_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
        if number == number.to_integral():
            return str(number.to_integral())
        return format_value(number)
    return str(value).strip()


def run_biff_extractor(workbook_path: Path) -> list[WorksheetRow]:
    ole = olefile.OleFileIO(str(workbook_path))
    try:
        workbook_stream = ole.openstream("Workbook").read()
    finally:
        ole.close()

    shared_strings = extract_shared_strings(workbook_stream)
    sheets = extract_sheet_info(workbook_stream)

    rows: list[WorksheetRow] = []
    for sheet in sheets:
        extracted_rows = parse_sheet_rows(workbook_stream, sheet["offset"], shared_strings)
        for row_index in sorted(extracted_rows):
            values = extracted_rows[row_index]
            cells = []
            for col_index in sorted(values):
                text = biff_value_to_text(values[col_index])
                if text:
                    cells.append(text)
            if cells:
                rows.append(
                    WorksheetRow(
                        sheet=str(sheet["name"]),
                        row_number=row_index + 1,
                        cells=cells,
                    )
                )

    if not rows:
        raise DomainParsingError("Nenhuma linha útil foi encontrada na planilha.")
    return rows


def run_powershell_extractor(workbook_path: Path) -> list[WorksheetRow]:
    workbook_name = workbook_path.name.replace("'", "''")
    command_text = f"""
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$ErrorActionPreference = "Stop"
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open((Join-Path (Get-Location) '{workbook_name}'))
    foreach ($worksheet in $workbook.Worksheets) {{
        $usedRange = $worksheet.UsedRange
        $rowCount = $usedRange.Rows.Count
        $colCount = $usedRange.Columns.Count
        for ($row = 1; $row -le $rowCount; $row++) {{
            $cells = New-Object System.Collections.Generic.List[string]
            for ($col = 1; $col -le $colCount; $col++) {{
                $text = [string]$worksheet.Cells.Item($row, $col).Text
                if (-not [string]::IsNullOrWhiteSpace($text)) {{
                    [void]$cells.Add($text.Trim())
                }}
            }}
            if ($cells.Count -gt 0) {{
                [PSCustomObject]@{{
                    sheet = [string]$worksheet.Name
                    row = $row
                    cells = $cells
                }} | ConvertTo-Json -Compress -Depth 4
            }}
        }}
    }}
}}
finally {{
    if ($workbook) {{
        $workbook.Close($false)
        [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
    }}
    if ($excel) {{
        $excel.Quit()
        [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
    }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    result = subprocess.run(
        [
            "powershell",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            command_text,
        ],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        cwd=str(workbook_path.parent),
    )

    rows: list[WorksheetRow] = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        payload = json.loads(line)
        rows.append(
            WorksheetRow(
                sheet=str(payload["sheet"]),
                row_number=int(payload["row"]),
                cells=[str(cell) for cell in payload["cells"]],
            )
        )
    if not rows:
        raise DomainParsingError("Nenhuma linha útil foi encontrada na planilha.")
    return rows


def extract_shared_strings(workbook_stream: bytes) -> list[str]:
    strings: list[str] = []
    position = 0
    while position < len(workbook_stream) - 4:
        record_type = struct.unpack("<H", workbook_stream[position : position + 2])[0]
        record_length = struct.unpack("<H", workbook_stream[position + 2 : position + 4])[0]
        record_data = workbook_stream[position + 4 : position + 4 + record_length]

        if record_type == 0x00FC:
            if len(record_data) < 8:
                break
            unique_strings = struct.unpack("<I", record_data[4:8])[0]
            sst_position = 8

            for _ in range(unique_strings):
                if sst_position + 3 > len(record_data):
                    break
                string_length = struct.unpack("<H", record_data[sst_position : sst_position + 2])[0]
                flags = record_data[sst_position + 2]
                sst_position += 3

                rich_text_runs = 0
                extension_size = 0
                if flags & 0x08:
                    rich_text_runs = struct.unpack("<H", record_data[sst_position : sst_position + 2])[0]
                    sst_position += 2
                if flags & 0x04:
                    extension_size = struct.unpack("<I", record_data[sst_position : sst_position + 4])[0]
                    sst_position += 4

                if flags & 0x01:
                    byte_length = string_length * 2
                    value = record_data[sst_position : sst_position + byte_length].decode(
                        "utf-16-le",
                        errors="replace",
                    )
                    sst_position += byte_length
                else:
                    value = record_data[sst_position : sst_position + string_length].decode(
                        "latin-1",
                        errors="replace",
                    )
                    sst_position += string_length

                sst_position += rich_text_runs * 4 + extension_size
                strings.append(value)
            break

        position += 4 + record_length
    return strings


def extract_sheet_info(workbook_stream: bytes) -> list[dict[str, int | str]]:
    sheets: list[dict[str, int | str]] = []
    position = 0
    while position < len(workbook_stream) - 4:
        record_type = struct.unpack("<H", workbook_stream[position : position + 2])[0]
        record_length = struct.unpack("<H", workbook_stream[position + 2 : position + 4])[0]
        record_data = workbook_stream[position + 4 : position + 4 + record_length]

        if record_type == 0x0085 and len(record_data) >= 8:
            offset = struct.unpack("<I", record_data[:4])[0]
            name_length = record_data[6]
            name_flag = record_data[7]
            if name_flag == 0:
                name = record_data[8 : 8 + name_length].decode("latin-1", errors="replace")
            else:
                name = record_data[8 : 8 + name_length * 2].decode("utf-16-le", errors="replace")
            sheets.append({"name": name, "offset": offset})

        position += 4 + record_length
    return sheets


def parse_sheet_rows(
    workbook_stream: bytes,
    offset: int,
    shared_strings: list[str],
) -> dict[int, dict[int, str | float]]:
    rows: dict[int, dict[int, str | float]] = {}
    position = offset
    in_sheet = False

    while position < len(workbook_stream) - 4:
        record_type = struct.unpack("<H", workbook_stream[position : position + 2])[0]
        record_length = struct.unpack("<H", workbook_stream[position + 2 : position + 4])[0]
        record_data = workbook_stream[position + 4 : position + 4 + record_length]

        if record_type == 0x0809:
            in_sheet = True
        elif record_type == 0x000A and in_sheet:
            break
        elif record_type == 0x00FD and len(record_data) >= 10:
            row_index = struct.unpack("<H", record_data[0:2])[0]
            col_index = struct.unpack("<H", record_data[2:4])[0]
            sst_index = struct.unpack("<I", record_data[6:10])[0]
            if sst_index < len(shared_strings):
                rows.setdefault(row_index, {})[col_index] = shared_strings[sst_index]
        elif record_type == 0x0203 and len(record_data) >= 14:
            row_index = struct.unpack("<H", record_data[0:2])[0]
            col_index = struct.unpack("<H", record_data[2:4])[0]
            value = struct.unpack("<d", record_data[6:14])[0]
            rows.setdefault(row_index, {})[col_index] = value
        elif record_type == 0x027E and len(record_data) >= 10:
            row_index = struct.unpack("<H", record_data[0:2])[0]
            col_index = struct.unpack("<H", record_data[2:4])[0]
            rows.setdefault(row_index, {})[col_index] = decode_rk_value(record_data[6:10])
        elif record_type == 0x00BD and len(record_data) >= 6:
            row_index = struct.unpack("<H", record_data[0:2])[0]
            first_col_index = struct.unpack("<H", record_data[2:4])[0]
            rk_position = 4
            col_index = first_col_index
            row_values = rows.setdefault(row_index, {})
            while rk_position + 6 <= len(record_data) - 2:
                row_values[col_index] = decode_rk_value(record_data[rk_position + 2 : rk_position + 6])
                col_index += 1
                rk_position += 6

        position += 4 + record_length

    return rows


def decode_rk_value(raw_value: bytes) -> float:
    rk_value = struct.unpack("<I", raw_value)[0]
    if rk_value & 0x02:
        value = rk_value >> 2
        if rk_value & 0x80000000:
            value -= 1 << 30
    else:
        double_bytes = struct.pack("<Q", (rk_value & 0xFFFFFFFC) << 32)
        value = struct.unpack("<d", double_bytes)[0]
    if rk_value & 0x01:
        value /= 100.0
    return float(value)


def biff_value_to_text(value: str | float) -> str:
    if isinstance(value, str):
        return value.strip()
    number = Decimal(str(value))
    if number == number.to_integral():
        return str(number.to_integral())
    return format_value(number)


def extract_company_code(rows: Iterable[WorksheetRow]) -> str:
    for row in rows:
        if len(row.cells) >= 2 and row.cells[0] == "Empresa:":
            match = re.match(r"^(\d+)\s*-", row.cells[1])
            if match:
                return match.group(1)
    raise DomainParsingError("Não foi possível identificar o código da filial na planilha.")


def split_page_one_blocks(rows: list[WorksheetRow]) -> list[CompetencyBlock]:
    blocks: list[CompetencyBlock] = []
    current_date: date | None = None
    current_rows: list[WorksheetRow] = []

    def flush_current() -> None:
        if not current_date or not current_rows:
            return
        joined = normalize_key("\n".join(" | ".join(item.cells) for item in current_rows))
        is_page_one = "FOLHA MENSAL" in joined or "LIQUIDO GERAL:" in joined
        if is_page_one:
            blocks.append(CompetencyBlock(competency_date=current_date, rows=current_rows.copy()))

    for row in rows:
        if len(row.cells) >= 2 and normalize_key(row.cells[0]) == "COMPETENCIA:":
            next_date = parse_competency_date(row.cells[1])
            if current_date is None:
                current_date = next_date
                current_rows = [row]
            elif next_date != current_date:
                flush_current()
                current_date = next_date
                current_rows = [row]
            else:
                current_rows.append(row)
        elif current_date is not None:
            current_rows.append(row)

    flush_current()
    if not blocks:
        raise DomainParsingError("Não foi possível localizar os blocos mensais da folha.")
    return blocks


def is_thirteenth_summary_workbook(rows: Iterable[WorksheetRow]) -> bool:
    for row in rows:
        if len(row.cells) >= 2 and normalize_key(row.cells[0]) == "CALCULO:":
            if "13" in normalize_key(row.cells[1]):
                return True
    return False


def parse_competency_date(raw_value: str) -> date:
    if "/" in raw_value:
        return datetime.strptime(raw_value, "%d/%m/%Y").date()
    serial = int(Decimal(raw_value))
    excel_epoch = date(1899, 12, 30)
    return date.fromordinal(excel_epoch.toordinal() + serial)


def special_thirteenth_posting_date(competency_date: date) -> str:
    return f"20/12/{competency_date.year}"


def find_value_after_label(block: CompetencyBlock, label: str) -> Decimal:
    normalized_label = normalize_key(label)
    for row in block.rows:
        for index, cell in enumerate(row.cells[:-1]):
            if normalize_key(cell) == normalized_label:
                try:
                    return parse_decimal(row.cells[index + 1])
                except Exception:
                    continue
    return Decimal("0.00")


def collect_rubrics(block: CompetencyBlock) -> dict[str, Decimal]:
    rubrics: dict[str, Decimal] = {}
    for row in block.rows:
        if len(row.cells) >= 4 and row.cells[0].isdigit():
            value = find_last_numeric_cell(row.cells[2:])
            if value is not None:
                rubrics[row.cells[1]] = value
    return rubrics


def collect_provento_rubrics(block: CompetencyBlock) -> dict[str, Decimal]:
    rubrics: dict[str, Decimal] = {}
    in_proventos = False
    for row in block.rows:
        if row.cells[0] == "PROVENTOS":
            in_proventos = True
            continue
        if row.cells[0] == "DESCONTOS":
            break
        if in_proventos and len(row.cells) >= 4 and row.cells[0].isdigit():
            value = find_last_numeric_cell(row.cells[2:])
            if value is not None:
                rubrics[row.cells[1]] = rubrics.get(row.cells[1], Decimal("0.00")) + value
    return rubrics


def is_vacation_provento(name: str) -> bool:
    normalized = normalize_key(name)
    if normalized in {
        "VANTAGENS ABONO",
        "MEDIA VALOR ABONO",
        "MEDIAS HORAS ABONO",
        "ESTORNO DESC PROV EMPRESTIMO TRAB FERIAS",
    }:
        return False
    return (
        "FERIAS" in normalized
        or " FER " in f" {normalized} "
        or "ABONO" in normalized
    )


def is_thirteenth_provento(name: str) -> bool:
    normalized = normalize_key(name)
    return "13" in normalized


def is_inss_discount_provento(name: str) -> bool:
    normalized = normalize_key(name)
    return normalized in {
        "SALARIO FAMILIA",
        "SALARIO FAMILIA RETROATIVO",
        "SALARIO MATERNIDADE DIAS",
        "ANUENIO LIC.MATERN",
    }


def find_values_after_labels(block: CompetencyBlock, labels: Iterable[str]) -> list[tuple[str, Decimal]]:
    normalized_labels = {normalize_key(label): sanitize_text(label).rstrip(":") for label in labels}
    values: list[tuple[str, Decimal]] = []
    for row in block.rows:
        for index, cell in enumerate(row.cells[:-1]):
            normalized_cell = normalize_key(cell)
            if normalized_cell not in normalized_labels:
                continue
            try:
                value = parse_decimal(row.cells[index + 1])
            except Exception:
                continue
            values.append((normalized_labels[normalized_cell], value))
    return values


def is_absence_discount_rubric(name: str) -> bool:
    normalized = normalize_key(name)
    return normalized in {
        "DESCONTO DIAS AFASTADOS",
        "DIAS FALTAS",
        "HORAS FALTAS PARCIAL",
        "DESC.VALOR PAGO A MAIOR",
    }


def is_employee_irrf_rubric(name: str) -> bool:
    normalized = normalize_key(name)
    return normalized in {
        "IMPOSTO DE RENDA",
        "IRRF FERIAS",
        "IRRF SOBRE RESCISAO",
        "IRRF 13O SALARIO RESCISAO",
    }


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.upper()


def find_last_numeric_cell(cells: list[str]) -> Decimal | None:
    for cell in reversed(cells):
        try:
            return parse_decimal(cell)
        except Exception:
            continue
    return None


def parse_decimal(raw_value: str | Decimal | None) -> Decimal:
    if raw_value is None:
        return Decimal("0.00")
    if isinstance(raw_value, Decimal):
        return raw_value.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    normalized = str(raw_value).replace('"', "").strip()
    if not normalized:
        return Decimal("0.00")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    return Decimal(normalized).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)


def decimal_to_storage(value: Decimal) -> str:
    return format(value.quantize(TWO_PLACES, rounding=ROUND_HALF_UP), "f")


def storage_to_decimal(value: object) -> Decimal:
    return Decimal(str(value)).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)


def month_reference(competency_date: date) -> str:
    return f"{competency_date.month:02d}/{competency_date.year}"


def month_end_display(competency_date: date) -> str:
    if competency_date.month == 12:
        next_month = date(competency_date.year + 1, 1, 1)
    else:
        next_month = date(competency_date.year, competency_date.month + 1, 1)
    final_date = date.fromordinal(next_month.toordinal() - 1)
    return f"{final_date.day}/{final_date.month}/{final_date.year}"


def format_value(value: Decimal) -> str:
    quantized = value.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    text = format(quantized, "f").replace(".", ",")
    if "," in text:
        text = text.rstrip("0").rstrip(",")
    return text


def sanitize_text(value: object) -> str:
    text = str(value)
    replacements = [
        ("Âº", "\u00ba"),
        ("Âª", "\u00aa"),
        ("繙", "\u00ba"),
        ("穠", "\u00aa"),
        ("L黠xadquido", "L\u00edquido"),
        ("l黠xadquido", "l\u00edquido"),
        ("LÃ­quido", "L\u00edquido"),
        ("lÃ­quido", "l\u00edquido"),
        ("CompetÃªncia", "Compet\u00eancia"),
        ("CompetÃƒÂªncia", "Compet\u00eancia"),
    ]
    for source, target in replacements:
        text = text.replace(source, target)
    return text


def source_label(source_key: str) -> str:
    return sanitize_text(SOURCE_DEFINITIONS.get(source_key, {}).get("label", source_key))


def build_context(block: CompetencyBlock, branch_code: str | None = None) -> CalculationContext:
    rubrics = collect_rubrics(block)
    provento_rubrics = collect_provento_rubrics(block)
    use_branch_67_rules = branch_code == "67"

    inss_discount_components = (
        [
            (name, value)
            for name, value in find_values_after_labels(
                block,
                ["(-) Salario Familia:", "(-) Salario Maternidade:"],
            )
            if value > 0
        ]
        if use_branch_67_rules
        else []
    )
    if not inss_discount_components and use_branch_67_rules:
        inss_discount_components = [
            (name, value)
            for name, value in provento_rubrics.items()
            if is_inss_discount_provento(name)
            if value > 0
        ]
    inss_difference_components = [
        (name, value)
        for name, value in rubrics.items()
        if normalize_key(name) == "INSS DIF FER DESC A MAIOR"
        if value > 0
    ] if use_branch_67_rules else []
    if inss_difference_components:
        inss_discount_components.extend(inss_difference_components)
    health_plan_reimbursement_components = [
        (name, value)
        for name, value in provento_rubrics.items()
        if normalize_key(name) == "REEMBOLSO PLANO DE SAUDE"
        if value > 0
    ] if use_branch_67_rules else []
    termination_charge_debit_names = {
        name
        for name, value in rubrics.items()
        if normalize_key(name) in {
            "MULTA ESTABILIDADE ART. 479/CLT",
            "ESTOURO RESCISAO",
            "AVISO PREVIO",
            "MEDIA VALOR AVISO PREVIO",
            "VANTAGENS AVISO PREVIO",
        }
        if value > 0
    } if use_branch_67_rules else set()
    inss_discount_provento_names = {
        name
        for name, value in provento_rubrics.items()
        if is_inss_discount_provento(name)
        if value > 0
    } if use_branch_67_rules else set()

    regular_components = [
        (name, value)
        for name, value in provento_rubrics.items()
        if not is_vacation_provento(name) and not is_thirteenth_provento(name)
        if name not in inss_discount_provento_names
        if name not in {item[0] for item in health_plan_reimbursement_components}
        if name not in termination_charge_debit_names
    ]
    vacation_components = [
        (name, value)
        for name, value in provento_rubrics.items()
        if is_vacation_provento(name)
        if not (
            use_branch_67_rules
            and normalize_key(name) == "INSS DIF FER DESC A MAIOR"
            and block.competency_date.month != 4
        )
        if value > 0
    ]
    if use_branch_67_rules and block.competency_date.month == 10:
        estorno_ferias = rubrics.get("ESTORNO DESC PROV EMPRESTIMO TRAB FERIAS", Decimal("0.00"))
        if estorno_ferias > 0:
            regular_components = [
                (name, value)
                for name, value in regular_components
                if normalize_key(name) != "ESTORNO DESC PROV EMPRESTIMO TRAB FERIAS"
            ]
            vacation_components.append(("ESTORNO DESC PROV EMPRESTIMO TRAB FERIAS", estorno_ferias))
    if use_branch_67_rules and block.competency_date.month == 4:
        inss_dif_ferias = rubrics.get("INSS DIF FER DESC A MAIOR", Decimal("0.00"))
        if inss_dif_ferias > 0:
            regular_components.append(("AJUSTE INSS DIF FER DESC A MAIOR", -inss_dif_ferias))
    thirteenth_components = [
        (name, value)
        for name, value in provento_rubrics.items()
        if is_thirteenth_provento(name)
        if value > 0
    ]
    vacation_advance_components = [
        ("ADIANTAMENTO DE FERIAS", rubrics.get("ADIANTAMENTO DE FERIAS", Decimal("0.00")))
    ]
    transport_discount_components = [
        ("VALE TRANSPORTE", rubrics.get("VALE TRANSPORTE", Decimal("0.00")))
    ]
    termination_net_components = [
        ("LIQUIDO RESCISAO", rubrics.get("LIQUIDO RESCISAO", Decimal("0.00")))
    ] if use_branch_67_rules else []
    termination_charge_debit_components = [
        (name, value)
        for name, value in rubrics.items()
        if normalize_key(name) in {
            "MULTA ESTABILIDADE ART. 479/CLT",
            "ESTOURO RESCISAO",
            "AVISO PREVIO",
            "MEDIA VALOR AVISO PREVIO",
            "VANTAGENS AVISO PREVIO",
        }
        if value > 0
    ] if use_branch_67_rules else []
    has_regular_prior_notice = any(
        normalize_key(name) in {
            "AVISO PREVIO",
            "MEDIA VALOR AVISO PREVIO",
            "VANTAGENS AVISO PREVIO",
        }
        and value > 0
        for name, value in rubrics.items()
    ) if use_branch_67_rules else False
    termination_charge_credit_components = [
        (name, value)
        for name, value in rubrics.items()
        if normalize_key(name) == "MULTA ESTABILIDADE ART. 480/CLT"
        if value > 0
    ] if use_branch_67_rules else []
    prior_notice_recovery_value = rubrics.get("AVISO PREVIO REAVIDO", Decimal("0.00"))
    prior_notice_recovery_components = (
        [("AVISO PREVIO REAVIDO", prior_notice_recovery_value)]
        if use_branch_67_rules and not has_regular_prior_notice and prior_notice_recovery_value > 0
        else []
    )
    absence_adjustment_components = [
        (name, value)
        for name, value in rubrics.items()
        if is_absence_discount_rubric(name)
        if value > 0
    ] if use_branch_67_rules else []
    if use_branch_67_rules and has_regular_prior_notice and prior_notice_recovery_value > 0:
        absence_adjustment_components.append(("AVISO PREVIO REAVIDO", prior_notice_recovery_value))
    employee_irrf_components = [
        (name, value)
        for name, value in rubrics.items()
        if is_employee_irrf_rubric(name)
        if value > 0
    ] if use_branch_67_rules else []
    health_plan_discount_components = [
        ("DESCONTO PLANO DE SAÚDE", rubrics.get("DESCONTO PLANO DE SAÚDE", Decimal("0.00")))
    ] if use_branch_67_rules else []
    meal_voucher_discount_components = [
        ("DESC VALE ALIMENTACAO", rubrics.get("DESC VALE ALIMENTACAO", Decimal("0.00")))
    ] if use_branch_67_rules else []
    alimony_components = [
        ("PENSAO ALIMENTICIA", rubrics.get("PENSAO ALIMENTICIA", Decimal("0.00")))
    ] if use_branch_67_rules else []
    thirteenth_rescisao_discount_components = [
        (name, value)
        for name, value in rubrics.items()
        if normalize_key(name) in {
            "ADIANTAMENTO 13 SALARIO RESCISAO",
            "DESCONTO DIFERENCA MEDIA HORA 13O",
            "DESCONTO DIFERENCA MEDIA VALOR 13O",
        }
        if value > 0
    ] if use_branch_67_rules else []
    payroll_discount_components = [
        (name, value)
        for name, value in rubrics.items()
        if (
            normalize_key(name).startswith("DESC. EMP. CRED. TRAB")
            or normalize_key(name).startswith("DESC EMP CRED TRAB FE")
            or normalize_key(name) in {
                "DESC.ADIANT.SALARIAL",
                "PROVISAO DESC. EMP. CRED. TRAB. FERIAS",
            }
        )
        if value > 0
    ] if use_branch_67_rules else []
    payroll_discount_components_for_absence = list(payroll_discount_components)
    if (
        use_branch_67_rules
        and (termination_charge_debit_components and termination_charge_credit_components)
    ):
        payroll_discount_components = []
    merged_absence_adjustment_components = list(absence_adjustment_components)
    payroll_discount_is_merged_into_absence = bool(
        use_branch_67_rules and absence_adjustment_components and payroll_discount_components_for_absence
    )
    if payroll_discount_is_merged_into_absence:
        merged_absence_adjustment_components.extend(payroll_discount_components_for_absence)
    employee_inss_components = [("Segurados", find_value_after_label(block, "Segurados:"))]
    if inss_difference_components:
        employee_inss_components.extend(inss_difference_components)
    fgts_components = [
        ("Valor do FGTS", find_value_after_label(block, "Valor do FGTS:")),
        ("Valor do FGTS Aprendiz", find_value_after_label(block, "Valor do FGTS Aprendiz:")),
    ]
    if not use_branch_67_rules:
        fgts_components.extend(
            [
                ("Valor FGTS Rescisorio", find_value_after_label(block, "Valor FGTS Rescisorio:")),
                ("Valor FGTS Resc. mes ant.", find_value_after_label(block, "Valor FGTS Resc. mes ant.:")),
            ]
        )
    pis_components = [("Valor PIS", find_value_after_label(block, "Valor PIS:"))]
    net_salary_components = [("Líquido Geral", find_value_after_label(block, "Líquido Geral:"))]
    employer_components = [
        ("Empresa", find_value_after_label(block, "Empresa:")),
        ("RAT", find_value_after_label(block, "RAT:")),
        ("Contribuintes", find_value_after_label(block, "Contribuintes:")),
        ("Terceiros", find_value_after_label(block, "Terceiros:")),
    ]

    values = {
        "vacation_total": sum_decimal(value for _, value in vacation_components),
        "thirteenth_difference": sum_decimal(value for _, value in thirteenth_components),
        "regular_salary": sum_decimal(value for _, value in regular_components),
        "inss_discounts": sum_decimal(value for _, value in inss_discount_components),
        "health_plan_reimbursement": sum_decimal(
            value for _, value in health_plan_reimbursement_components
        ),
        "vacation_advance": rubrics.get("ADIANTAMENTO DE FERIAS", Decimal("0.00")),
        "transport_discount": transport_discount_components[0][1],
        "termination_net": termination_net_components[0][1] if termination_net_components else Decimal("0.00"),
        "termination_charge_debit": sum_decimal(value for _, value in termination_charge_debit_components),
        "termination_charge_credit": sum_decimal(value for _, value in termination_charge_credit_components),
        "prior_notice_recovery": sum_decimal(value for _, value in prior_notice_recovery_components),
        "absence_adjustment": sum_decimal(value for _, value in merged_absence_adjustment_components),
        "net_salary": net_salary_components[0][1],
        "employee_irrf": sum_decimal(value for _, value in employee_irrf_components),
        "employee_inss": sum_decimal(value for _, value in employee_inss_components),
        "fgts": sum_decimal(value for _, value in fgts_components),
        "pis": pis_components[0][1],
        "health_plan_discount": health_plan_discount_components[0][1] if health_plan_discount_components else Decimal("0.00"),
        "meal_voucher_discount": meal_voucher_discount_components[0][1] if meal_voucher_discount_components else Decimal("0.00"),
        "alimony": alimony_components[0][1] if alimony_components else Decimal("0.00"),
        "thirteenth_rescisao_discount": sum_decimal(
            value for _, value in thirteenth_rescisao_discount_components
        ),
        "payroll_discount": (
            Decimal("0.00")
            if payroll_discount_is_merged_into_absence
            else sum_decimal(value for _, value in payroll_discount_components)
        ),
        "employer_inss": sum_decimal(value for _, value in employer_components),
    }
    components = {
        "vacation_total": vacation_components,
        "thirteenth_difference": thirteenth_components,
        "regular_salary": regular_components,
        "inss_discounts": inss_discount_components,
        "health_plan_reimbursement": health_plan_reimbursement_components,
        "vacation_advance": [
            item for item in vacation_advance_components if item[1] > 0
        ],
        "transport_discount": [
            item for item in transport_discount_components if item[1] > 0
        ],
        "termination_net": [item for item in termination_net_components if item[1] > 0],
        "termination_charge_debit": termination_charge_debit_components,
        "termination_charge_credit": termination_charge_credit_components,
        "prior_notice_recovery": [item for item in prior_notice_recovery_components if item[1] > 0],
        "absence_adjustment": merged_absence_adjustment_components,
        "net_salary": [item for item in net_salary_components if item[1] > 0],
        "employee_irrf": employee_irrf_components,
        "employee_inss": [item for item in employee_inss_components if item[1] > 0],
        "fgts": [item for item in fgts_components if item[1] > 0],
        "pis": [item for item in pis_components if item[1] > 0],
        "health_plan_discount": [
            item for item in health_plan_discount_components if item[1] > 0
        ],
        "meal_voucher_discount": [
            item for item in meal_voucher_discount_components if item[1] > 0
        ],
        "alimony": [item for item in alimony_components if item[1] > 0],
        "thirteenth_rescisao_discount": [
            item for item in thirteenth_rescisao_discount_components if item[1] > 0
        ],
        "payroll_discount": (
            []
            if payroll_discount_is_merged_into_absence
            else [item for item in payroll_discount_components if item[1] > 0]
        ),
        "employer_inss": [item for item in employer_components if item[1] > 0],
    }
    flags = {
        "has_special_opening": values["vacation_total"] > 0 or values["thirteenth_difference"] > 0
    }

    return CalculationContext(
        competency_date=block.competency_date,
        posting_date=month_end_display(block.competency_date),
        reference=month_reference(block.competency_date),
        values=values,
        components=components,
        flags=flags,
    )


def build_thirteenth_summary_context(rows: list[WorksheetRow]) -> CalculationContext:
    competency_row = next(
        (
            row
            for row in rows
            if len(row.cells) >= 2 and row.cells[0] == "CompetÃªncia:"
        ),
        None,
    )
    if competency_row is None:
        competency_row = next(
            (
                row
                for row in rows
                if len(row.cells) >= 2 and normalize_key(row.cells[0]) == "COMPETENCIA:"
            ),
            None,
        )
    if competency_row is None:
        raise DomainParsingError("NÃ£o foi possÃ­vel identificar a competÃªncia do arquivo do 13Âº.")

    competency_date = parse_competency_date(competency_row.cells[1])
    start_index = rows.index(competency_row)
    block = CompetencyBlock(competency_date=competency_date, rows=rows[start_index:])
    rubrics = collect_rubrics(block)
    provento_rubrics = collect_provento_rubrics(block)

    thirteenth_components = [
        (name, value)
        for name, value in provento_rubrics.items()
        if is_thirteenth_provento(name)
        if value > 0
    ]
    thirteenth_advance_components = [
        ("ADIANTAMENTO 13 SALARIO", rubrics.get("ADIANTAMENTO 13 SALARIO", Decimal("0.00")))
    ]
    thirteenth_employee_inss_components = [("Segurados", find_value_after_label(block, "Segurados:"))]
    thirteenth_inss_discount_components = find_values_after_labels(
        block,
        ("(-) Salário Família:", "(-) Salário Maternidade:"),
    )
    thirteenth_irrf_components = [
        ("Valor IRRF 13º Salário", find_value_after_label(block, "Valor IRRF 13º Salário:"))
    ]
    thirteenth_fgts_components = [("Valor do FGTS", find_value_after_label(block, "Valor do FGTS:"))]
    thirteenth_pis_components = [("Valor PIS", find_value_after_label(block, "Valor PIS:"))]
    thirteenth_net_components = [("LÃ­quido Geral", find_value_after_label(block, "LÃ­quido Geral:"))]
    thirteenth_net_components = [("Líquido Geral", find_value_after_label(block, "Líquido Geral:"))]
    thirteenth_employer_components = [
        ("Empresa", find_value_after_label(block, "Empresa:")),
        ("RAT", find_value_after_label(block, "RAT:")),
        ("Terceiros", find_value_after_label(block, "Terceiros:")),
    ]

    values = {
        "vacation_total": Decimal("0.00"),
        "thirteenth_difference": Decimal("0.00"),
        "regular_salary": Decimal("0.00"),
        "vacation_advance": Decimal("0.00"),
        "transport_discount": Decimal("0.00"),
        "net_salary": Decimal("0.00"),
        "employee_inss": Decimal("0.00"),
        "fgts": Decimal("0.00"),
        "pis": Decimal("0.00"),
        "employer_inss": Decimal("0.00"),
        "thirteenth_salary": sum_decimal(value for _, value in thirteenth_components),
        "thirteenth_net_salary": thirteenth_net_components[0][1],
        "thirteenth_advance": thirteenth_advance_components[0][1],
        "thirteenth_employee_inss": thirteenth_employee_inss_components[0][1],
        "thirteenth_inss_discounts": sum_decimal(
            value for _, value in thirteenth_inss_discount_components
        ),
        "thirteenth_irrf": thirteenth_irrf_components[0][1],
        "thirteenth_fgts": thirteenth_fgts_components[0][1],
        "thirteenth_pis": thirteenth_pis_components[0][1],
        "thirteenth_employer_inss": sum_decimal(value for _, value in thirteenth_employer_components),
    }
    components = {
        "vacation_total": [],
        "thirteenth_difference": [],
        "regular_salary": [],
        "vacation_advance": [],
        "transport_discount": [],
        "net_salary": [],
        "employee_inss": [],
        "fgts": [],
        "pis": [],
        "employer_inss": [],
        "thirteenth_salary": thirteenth_components,
        "thirteenth_net_salary": [item for item in thirteenth_net_components if item[1] > 0],
        "thirteenth_advance": [item for item in thirteenth_advance_components if item[1] > 0],
        "thirteenth_employee_inss": [item for item in thirteenth_employee_inss_components if item[1] > 0],
        "thirteenth_inss_discounts": [
            item for item in thirteenth_inss_discount_components if item[1] > 0
        ],
        "thirteenth_irrf": [item for item in thirteenth_irrf_components if item[1] > 0],
        "thirteenth_fgts": [item for item in thirteenth_fgts_components if item[1] > 0],
        "thirteenth_pis": [item for item in thirteenth_pis_components if item[1] > 0],
        "thirteenth_employer_inss": [item for item in thirteenth_employer_components if item[1] > 0],
    }
    flags = {"has_special_opening": values["thirteenth_salary"] > 0}

    return CalculationContext(
        competency_date=competency_date,
        posting_date=special_thirteenth_posting_date(competency_date),
        reference=month_reference(competency_date),
        values=values,
        components=components,
        flags=flags,
    )


def sum_decimal(values: Iterable[Decimal]) -> Decimal:
    total = Decimal("0.00")
    for value in values:
        total += value
    return total.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)


def apply_start_lot_strategy(rule: MappingRule, context: CalculationContext) -> bool:
    if rule.start_lot_strategy == "always":
        return True
    if rule.start_lot_strategy == "never":
        return False
    if rule.start_lot_strategy == "debit":
        return bool(rule.debit_account.strip())
    if rule.start_lot_strategy == "no_prior_specials":
        return not context.flags.get("has_special_opening", False)
    return False


def build_memory_text(context: CalculationContext, source_key: str) -> str:
    value = context.values.get(source_key, Decimal("0.00"))
    components = [item for item in context.components.get(source_key, []) if item[1] > 0]
    prefix = sanitize_text(
        SOURCE_DEFINITIONS.get(source_key, {}).get("memory_prefix", source_label(source_key))
    )

    if not components:
        return f"{prefix}: {format_value(value)}"
    if len(components) == 1:
        label, component_value = components[0]
        return f"{prefix}: {sanitize_text(label)} = {format_value(component_value)}"

    parts = " + ".join(
        f"{sanitize_text(label)} {format_value(component_value)}"
        for label, component_value in components
    )
    return f"{prefix}: {parts} = {format_value(value)}"


def render_history_template(rule: MappingRule, context: CalculationContext, branch_code: str) -> str:
    if (
        branch_code == "67"
        and rule.source_key == "termination_charge_debit"
        and context.reference == "04/2025"
    ):
        return "RESCISAO 04/2025"
    if branch_code == "67" and rule.source_key == "prior_notice_recovery":
        if context.reference == "03/2025":
            return "AVISO PREVIO 03/2025"
        return f"MULTA RESCISAO {context.reference}"
    payload = SafeFormatDict(
        reference=context.reference,
        branch_code=branch_code,
        month=f"{context.competency_date.month:02d}",
        year=str(context.competency_date.year),
        posting_date=context.posting_date,
        source_label=source_label(rule.source_key),
    )
    return sanitize_text(rule.history_template.format_map(payload))


def build_entries(
    contexts: list[CalculationContext],
    rules: list[MappingRule],
    branch_code: str,
) -> list[GeneratedEntry]:
    ordered_rules = sorted(rules, key=lambda item: (item.order, item.label.lower(), item.rule_id))
    entries: list[GeneratedEntry] = []

    for context in contexts:
        for rule in ordered_rules:
            if not rule.active:
                continue
            value = context.values.get(rule.source_key, Decimal("0.00"))
            if value <= 0:
                continue
            entries.append(
                GeneratedEntry(
                    competency_reference=context.reference,
                    mapping_label=rule.label,
                    source_key=rule.source_key,
                    source_label=source_label(rule.source_key),
                    posting_date=context.posting_date,
                    debit_account=rule.debit_account.strip(),
                    credit_account=rule.credit_account.strip(),
                    value=value,
                    history=render_history_template(rule, context, branch_code),
                    start_lot=apply_start_lot_strategy(rule, context),
                    memory=build_memory_text(context, rule.source_key),
                    branch_code=branch_code,
                    order=rule.order,
                )
            )

    return entries


def analysis_to_state(
    *,
    branch_code: str,
    workbook_name: str,
    contexts: list[CalculationContext],
    mapping_rules: list[MappingRule] | None = None,
    include_thirteenth_summary: bool = False,
) -> dict[str, object]:
    rules = (
        mapping_rules
        if mapping_rules is not None
        else default_mapping_rules(
            branch_code=branch_code,
            include_thirteenth_summary=include_thirteenth_summary,
        )
    )
    return {
        "branch_code": branch_code,
        "workbook_name": workbook_name,
        "contexts": [context.to_state() for context in contexts],
        "mapping_rules": [rule.to_dict() for rule in rules],
    }


def analyze_workbook(workbook_path: str | Path) -> dict[str, object]:
    workbook_path = Path(workbook_path)
    rows = run_excel_extractor(workbook_path)
    branch_code = extract_company_code(rows)
    if is_thirteenth_summary_workbook(rows):
        contexts = [build_thirteenth_summary_context(rows)]
        include_thirteenth_summary = True
    else:
        blocks = split_page_one_blocks(rows)
        contexts = [build_context(block, branch_code=branch_code) for block in blocks]
        include_thirteenth_summary = False
    return analysis_to_state(
        branch_code=branch_code,
        workbook_name=workbook_path.name,
        contexts=contexts,
        include_thirteenth_summary=include_thirteenth_summary,
    )


def merge_analysis_states(states: list[dict[str, object]]) -> dict[str, object]:
    if not states:
        raise DomainParsingError("Nenhum arquivo foi enviado para consolidaÃ§Ã£o.")

    branch_codes = {str(state["branch_code"]) for state in states}
    if len(branch_codes) != 1:
        raise DomainParsingError("Os arquivos enviados precisam pertencer Ã  mesma filial.")

    contexts: list[CalculationContext] = []
    include_thirteenth_summary = False
    workbook_names: list[str] = []

    for state in states:
        contexts.extend(contexts_from_state(state))
        workbook_names.append(str(state["workbook_name"]))
        if any(
            rule.source_key in THIRTEENTH_SUMMARY_SOURCE_KEYS
            for rule in mapping_rules_from_state(state)
        ):
            include_thirteenth_summary = True

    return analysis_to_state(
        branch_code=branch_codes.pop(),
        workbook_name=" + ".join(workbook_names),
        contexts=contexts,
        include_thirteenth_summary=include_thirteenth_summary,
    )


def contexts_from_state(state: dict[str, object]) -> list[CalculationContext]:
    return [CalculationContext.from_state(item) for item in list(state["contexts"])]


def mapping_rules_from_state(state: dict[str, object]) -> list[MappingRule]:
    rules = [MappingRule.from_dict(item) for item in list(state["mapping_rules"])]
    return apply_history_template_conventions(rules)


def source_options() -> list[dict[str, str | int]]:
    options = []
    for key, config in sorted(
        SOURCE_DEFINITIONS.items(),
        key=lambda item: (item[1]["order"], item[1]["label"]),
    ):
        options.append(
            {
                "key": key,
                "label": sanitize_text(config["label"]),
                "description": sanitize_text(config["description"]),
                "order": int(config["order"]),
            }
        )
    return options


def build_preview(state: dict[str, object]) -> dict[str, object]:
    branch_code = str(state["branch_code"])
    contexts = contexts_from_state(state)
    rules = mapping_rules_from_state(state)
    entries = build_entries(contexts, rules, branch_code)

    total_value = sum_decimal(entry.value for entry in entries)
    competencies = []
    for context in contexts:
        sources = []
        for option in source_options():
            source_key = str(option["key"])
            value = context.values.get(source_key, Decimal("0.00"))
            if value <= 0:
                continue
            sources.append(
                {
                    "key": source_key,
                    "label": str(option["label"]),
                    "description": str(option["description"]),
                    "value": format_value(value),
                    "memory": build_memory_text(context, source_key),
                }
            )
        competencies.append(
            {
                "reference": context.reference,
                "posting_date": context.posting_date,
                "sources": sources,
            }
        )

    preview_entries = []
    for index, entry in enumerate(entries, start=1):
        preview_entries.append(
            {
                "index": index,
                "competency_reference": entry.competency_reference,
                "mapping_label": entry.mapping_label,
                "source_label": entry.source_label,
                "posting_date": entry.posting_date,
                "debit_account": entry.debit_account,
                "credit_account": entry.credit_account,
                "value": format_value(entry.value),
                "history": entry.history,
                "start_lot": "1" if entry.start_lot else "",
                "memory": entry.memory,
            }
        )

    separator_before_entry_index = next(
        (
            index
            for index, entry in enumerate(entries)
            if entry.source_key in THIRTEENTH_SUMMARY_SOURCE_KEYS
        ),
        None,
    )
    entry_csv_rows = [entry.to_csv_row() for entry in entries]
    if separator_before_entry_index not in (None, 0):
        entry_csv_rows.insert(
            separator_before_entry_index,
            {field: "" for field in CSV_HEADER},
        )

    preview_rules = []
    for rule in sorted(rules, key=lambda item: (item.order, item.label.lower(), item.rule_id)):
        preview_rules.append(
            {
                **rule.to_dict(),
                "source_label": source_label(rule.source_key),
                "start_lot_label": START_LOT_OPTIONS.get(rule.start_lot_strategy, rule.start_lot_strategy),
            }
        )

    has_thirteenth_summary = any(
        any(
            context.values.get(source_key, Decimal("0.00")) > 0
            for source_key in THIRTEENTH_SUMMARY_SOURCE_KEYS
        )
        for context in contexts
    )
    needs_thirteenth_summary = (
        not has_thirteenth_summary
        and any(
            context.values.get("thirteenth_difference", Decimal("0.00")) > 0
            or context.values.get("thirteenth_rescisao_discount", Decimal("0.00")) > 0
            for context in contexts
        )
    )

    return {
        "branch_code": branch_code,
        "workbook_name": str(state["workbook_name"]),
        "has_thirteenth_summary": has_thirteenth_summary,
        "needs_thirteenth_summary": needs_thirteenth_summary,
        "separator_before_entry_index": separator_before_entry_index,
        "entries": preview_entries,
        "entry_csv_rows": entry_csv_rows,
        "mapping_rules": preview_rules,
        "source_options": source_options(),
        "competencies": competencies,
        "summary": {
            "competencies": len(contexts),
            "entries": len(entries),
            "total_value": format_value(total_value),
        },
    }


def generate_csv_rows(
    workbook_path: str | Path,
    supplementary_workbook_path: str | Path | None = None,
) -> tuple[str, list[dict[str, str]]]:
    state = analyze_workbook(workbook_path)
    if supplementary_workbook_path is not None:
        supplementary_state = analyze_workbook(supplementary_workbook_path)
        state = merge_analysis_states([state, supplementary_state])
    preview = build_preview(state)
    return str(preview["branch_code"]), list(preview["entry_csv_rows"])


def serialize_preview_for_excel(preview: dict[str, object]) -> bytes:
    workbook = Workbook()
    launches = workbook.active
    launches.title = "Lancamentos"
    launches.append(CSV_HEADER + ["Mapeamento", "Fonte", "Memória de cálculo"])
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in launches[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    separator_before_entry_index = preview.get("separator_before_entry_index")
    for index, entry in enumerate(list(preview["entries"])):
        if separator_before_entry_index == index:
            launches.append([""] * 13)
        launches.append(
            [
                entry["posting_date"],
                entry["debit_account"],
                entry["credit_account"],
                entry["value"],
                "",
                entry["history"],
                entry["start_lot"],
                preview["branch_code"],
                "",
                "",
                entry["mapping_label"],
                entry["source_label"],
                entry["memory"],
            ]
        )

    launches.freeze_panes = "A2"
    launches.auto_filter.ref = launches.dimensions

    memory_sheet = workbook.create_sheet("Memoria")
    memory_sheet.append(["Competência", "Data lançamento", "Fonte", "Valor", "Memória de cálculo"])
    for cell in memory_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for competency in list(preview["competencies"]):
        for source in list(competency["sources"]):
            memory_sheet.append(
                [
                    competency["reference"],
                    competency["posting_date"],
                    source["label"],
                    source["value"],
                    source["memory"],
                ]
            )
    memory_sheet.freeze_panes = "A2"
    memory_sheet.auto_filter.ref = memory_sheet.dimensions

    mapping_sheet = workbook.create_sheet("Mapeamento")
    mapping_sheet.append(
        [
            "Ativo",
            "Ordem",
            "Rótulo",
            "Fonte",
            "Conta débito",
            "Conta crédito",
            "Histórico",
            "Estratégia lote",
        ]
    )
    for cell in mapping_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for rule in list(preview["mapping_rules"]):
        mapping_sheet.append(
            [
                "Sim" if rule["active"] else "Não",
                rule["order"],
                rule["label"],
                rule["source_label"],
                rule["debit_account"],
                rule["credit_account"],
                rule["history_template"],
                START_LOT_OPTIONS.get(str(rule["start_lot_strategy"]), str(rule["start_lot_strategy"])),
            ]
        )
    mapping_sheet.freeze_panes = "A2"
    mapping_sheet.auto_filter.ref = mapping_sheet.dimensions

    for worksheet in workbook.worksheets:
        adjust_column_widths(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def adjust_column_widths(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 44)


def write_csv_bytes(csv_rows: list[dict[str, str]]) -> bytes:
    with TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / "dominio_lancamento.csv"
        with temp_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=CSV_HEADER)
            writer.writeheader()
            writer.writerows(csv_rows)
        return temp_path.read_bytes()
